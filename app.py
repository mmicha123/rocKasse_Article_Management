import locale
import plistlib
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import simpledialog
from tkinter import messagebox
from tkinter import OptionMenu
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl import Workbook

# globals
artikel_Data_Path = ""
artikel_Sparten = []
artikel_Data = []
artikel_Entrys = []
last_index = 0

locale.setlocale(locale.LC_ALL, "de_DE.UTF-8")  # German

# öfne APP
root = tk.Tk()
root.geometry("960x960")
root.title("Kassen Center")
root.resizable(0, 0)

canvas = tk.Canvas(root, borderwidth=0)
frame = tk.Frame(canvas)
vsb = tk.Scrollbar(root, orient="vertical", command=canvas.yview)
canvas.configure(yscrollcommand=vsb.set)

vsb.pack(side="right", fill="y")
canvas.pack(side="left", fill="both", expand=True)
canvas.create_window((0, 0), window=frame, anchor="nw")


def onFrameConfigure(canvas):
    canvas.configure(scrollregion=canvas.bbox("all"))

def _on_mousewheel(event):
    canvas.yview_scroll(int(-1*(event.delta/120)), "units")

frame.bind("<Configure>", lambda event, canvas=canvas: onFrameConfigure(canvas))
canvas.bind_all("<MouseWheel>", _on_mousewheel)

def valdidateFloat(action, value_if_allowed):
    if action == "1":
        try:
            loc_float = locale.atof(value_if_allowed)
            loc_float_format = locale.format_string("%f", loc_float)
            try:
                loc_same_length = loc_float_format[: len(value_if_allowed)]
                return value_if_allowed == loc_same_length
            except:
                return False
        except:
            return False
    else:
        return True


vcmd = (root.register(valdidateFloat), "%d", "%P")


class MyDialog:
    def __init__(self, parent):
        top = self.top = tk.Toplevel(parent)

        ttk.Label(top, text="Sparte").grid(column=0, row=0, sticky=tk.W)
        ttk.Label(top, text="ID").grid(column=1, row=0, sticky=tk.W)
        ttk.Label(top, text="Name").grid(column=2, row=0, sticky=tk.W)
        ttk.Label(top, text="Preis").grid(column=3, row=0, sticky=tk.W)
        ttk.Label(top, text="MWSt").grid(column=4, row=0, sticky=tk.W)

        self.var = tk.StringVar(top)
        self.var.set(artikel_Sparten[0])
        self.om = OptionMenu(top, self.var, *artikel_Sparten)
        self.om.grid(column=0, row=1, sticky=tk.W)

        self.id_entry = ttk.Entry(top, width=10, validate="key", validatecommand=vcmd)
        self.id_entry.grid(column=1, row=1, sticky=tk.W)

        self.name_entry = ttk.Entry(top)
        self.name_entry.grid(column=2, row=1, sticky=tk.W)

        self.preis_entry = ttk.Entry(top, validate="key", validatecommand=vcmd)
        self.preis_entry.grid(column=3, row=1, sticky=tk.W)

        self.mwst_entry = ttk.Entry(top, validate="key", validatecommand=vcmd)
        self.mwst_entry.grid(column=4, row=1, sticky=tk.W)

        self.submit_button = ttk.Button(top, text="Speichern", command=self.send)
        self.submit_button.grid(column=3, columnspan=2, row=2, sticky=tk.W)

    def send(self):
        self.id = int(self.id_entry.get())
        self.sparte = self.var.get()
        self.name = self.name_entry.get()
        self.preis = self.preis_entry.get()
        self.mwst = self.mwst_entry.get()

        if not self.id or not self.name or not self.preis or not self.mwst:
            self.id = None
            self.name = None
            self.preis = None
            self.mwst = None
            messagebox.showerror("Eingabe Fehler", "Bitte gib alle Felder an")
            self.top.focus()
            return

        test_id = [artikel_Data.index(tupl) for tupl in artikel_Data if tupl[0] == self.id]
        if test_id:
            messagebox.showerror("ID Fehler", "Bitte gib eine andere ID ein")
            self.top.focus()
            return
        
        create_new_artikle_call(self.sparte, self.id, self.name, self.preis, self.mwst)
        self.top.destroy()


def poulate_info():
    ttk.Label(frame, text="ID").grid(column=0, row=1, sticky=tk.W, padx=5, pady=5)
    ttk.Label(frame, text="Artikel").grid(column=1, row=1, sticky=tk.W, padx=5, pady=5)
    ttk.Label(frame, text="Preis").grid(column=2, row=1, sticky=tk.W, padx=5, pady=5)
    ttk.Label(frame, text="Preis AH").grid(column=3, row=1, sticky=tk.W, padx=5, pady=5)
    ttk.Label(frame, text="MWSt in %").grid(column=4, row=1, sticky=tk.W, padx=5, pady=5)
    ttk.Label(frame, text="MWSt AH in %").grid(column=5, row=1, sticky=tk.W, padx=5, pady=5)
    ttk.Label(frame, text="löschen?").grid(column=6, row=1, sticky=tk.W, padx=5, pady=5)

def poulate_ui():
    if not artikel_Data:
        return

    if artikel_Entrys:
        for widget in frame.winfo_children():
            widget.destroy()
        
        artikel_Sparten.clear()
        artikel_Entrys.clear()
        globals()["last_index"] = 0
        shift = 2
        poulate_info()

    shift = 2

    
    for i in range(len(artikel_Data)):
        (id, name, preis, preisah, mwst, mwstah, sparte, sparte_id, new_item) = artikel_Data[i]

        row_i = i + shift

        if sparte not in artikel_Sparten:
                artikel_Sparten.insert(sparte_id, sparte)
                ttk.Label(frame, text=sparte, foreground="#000ff0").grid(
                    columnspan=6, row=row_i, sticky=tk.W
                )
                shift = shift + 1
                row_i = row_i + 1

        ttk.Label(frame, text=id).grid(
            column=0, row=row_i, sticky=tk.W
        )
        artikel_lable = ttk.Label(frame, text=name)
        artikel_lable.grid(column=1, row=row_i, sticky=tk.W)
        if(new_item):
            artikel_lable.config(foreground="#00af00")

        preis_entry = ttk.Entry(
            frame, justify="right", width=10, validate="key", validatecommand=vcmd
        )
        preis_entry.grid(column=2, row=row_i, sticky=tk.W, padx=2)
        preis_entry.insert(0, preis)

        preisah_entry = ttk.Entry(
            frame, justify="right", width=10, validate="key", validatecommand=vcmd
        )
        preisah_entry.grid(column=3, row=row_i, sticky=tk.W, padx=2)
        preisah_entry.insert(0, preisah)

        mwst_entry = ttk.Entry(
            frame, justify="right", width=10, validate="key", validatecommand=vcmd
        )
        mwst_entry.grid(column=4, row=row_i, sticky=tk.W, padx=2)
        mwst_entry.insert(0, mwst)

        mwstah_entry = ttk.Entry(
            frame, justify="right", width=10, validate="key", validatecommand=vcmd
        )
        mwstah_entry.grid(column=5, row=row_i, sticky=tk.W, padx=2)
        mwstah_entry.insert(0, mwstah)

        delete_check = ttk.Checkbutton(frame)
        delete_check.grid(column=6, row=row_i, sticky=tk.W, padx=2)
        delete_check.state(["!alternate"])

        globals()["last_index"] = row_i

        artikel_Entrys.insert(i, (id, preis_entry, preisah_entry, mwst_entry, mwstah_entry, delete_check))

def select_File():
    artikel_Data.clear()

    globals()["artikel_Data_Path"] = filedialog.askopenfilename(
        initialdir="./", filetypes=[("PLIST-Datei", ".plist"), ("Alle Typen", "*.*")]
    )

    with open(artikel_Data_Path, "rb") as infile:
        fullArtikelList = plistlib.load(infile)

    for i in range(len(fullArtikelList)):
        artikel = fullArtikelList[i]

        sparte = artikel["Sparte"]
        sparte_id = int(artikel["SpartenId"])

        id = int(artikel["ArtikelId"])
        name = artikel["Artikel"]
        preis = artikel["Preis"]
        preisah = artikel["PreisAH"]
        mwst = artikel["MWSt"]
        mwstah = artikel["MWStAH"]

        artikel_Data.insert(i, (id, name, preis, preisah, mwst, mwstah, sparte, sparte_id, False))

    poulate_ui()

    infile.close()

def save_Data():
    if not artikel_Entrys:
        return
    
    tmp_artikel_Data = artikel_Data.copy()

    artikel_Data.clear()
    
    for i in range(len(artikel_Entrys)):
        (id_enty, preis_entry, preisah_entry, mwst_entry, mwstah_entry, delete_check) = artikel_Entrys[i]
        (id, name, preis, preisah, mwst, mwstah, sparte, sparte_id, new_item) = tmp_artikel_Data[i]

        if "selected" in delete_check.state():
            continue

        artikel_Data.insert(i, (id, name, preis_entry.get(), preisah_entry.get(), mwst_entry.get(), mwstah_entry.get(), sparte, sparte_id, new_item))

    poulate_ui()

def create_new_artikle_widget():
    if not artikel_Data:
        return
    dialog = MyDialog(root)
    root.wait_window(dialog.top)

def create_new_artikle_call(sparte, id, name, preis, mwst):
    sparteId = artikel_Sparten.index(sparte) + 1

    artikel_Data.append((id, name, preis, preis, mwst, mwst, sparte, sparteId, True))

    artikel_Data.sort(key=lambda x: x[0])
    
    poulate_ui()

def save_File():
    if not artikel_Data:
        return

    f = open(artikel_Data_Path, "rb")
    fullArtikelList = plistlib.load(f)

    #deletion and normal save of data
    for i in range(len(fullArtikelList) - 1, -1, -1):
        artikel_pl = fullArtikelList[i]

        #check for deletion
        artikel_pl_id = int(artikel_pl["ArtikelId"])
        a_d = [artikel_Data.index(tupl) for tupl in artikel_Data if tupl[0] == artikel_pl_id]
        
        #löschen von element
        if len(a_d) <= 0:
            del fullArtikelList[i]
            continue
        
        a_d_index = a_d[0]

        (id, name, preis, preisah, mwst, mwstah, sparte, sparte_id, new_item) = artikel_Data[a_d_index]

        if artikel_pl["Artikel"] == name:
            artikel_pl["Preis"] = preis
            artikel_pl["PreisAH"] = preisah
            artikel_pl["MWSt"] = mwst
            artikel_pl["MWStAH"] = mwstah
            
        #artikel_pl["Sparte"]
        #int(artikel_pl["SpartenId"])

    newList = []

    for i in range(len(artikel_Data)):
        (id, name, preis, preisah, mwst, mwstah, sparte, sparte_id, new_item) = artikel_Data[i]

        if not new_item:
            continue

        artikel = dict(
            Artikel = name,
            ArtikelId = str(id),
            BarCode = "",
            ComboArtikel = "",
            Drucker = "Theke",
            FreierPreis = "0",
            Gutschrift = "0",
            LongText = "",
            MWSt = str(mwst),
            MWStAH = str(mwstah),
            Maincourse = "0",
            NichtUmsatzRelevant = "1",
            PicturePath = "",
            Preis = str(preis),
            PreisAH = str(preisah),
            Preisliste = "Standard",
            RabattAusnahme = "",
            Rezeptur = "|",
            Sparte = sparte,
            SpartenId = str(sparte_id),
            SpartenSteuerung = "0",
            WG = "Speisen",
            Zusatztext = "",
            articleTSE = "Umsatz"
        )

        new_index = 0
        clostest_id = int(fullArtikelList[0]["ArtikelId"])

        for j in range(1, len(fullArtikelList)):
            artikel_id = int(fullArtikelList[j]["ArtikelId"])
            if abs(artikel_id - id) < abs(clostest_id - id):
                clostest_id = artikel_id
                new_index = j
        
        newList.append((new_index + 1, artikel))

    for i in range(len(newList)):
        (index, artikel) = newList[i]
        fullArtikelList.insert(index, artikel)

    f.close()

    f2 = open(artikel_Data_Path, "wb+")
    plistlib.dump(fullArtikelList, f2)

def save_to_excel():
    if not artikel_Data:
        return

    # entscheide ob von Plist zu Excel oder umgekehrt Convertiert

    # erstellt aka überschreibt die Exccel Datei
    wb = Workbook()
    ws = wb.create_sheet("Kasse", 0)

    # Überschriften
    ws["A1"].value = "ID"
    ws["B1"].value = "Artikel"
    ws["C1"].value = "Preis"
    ws["D1"].value = "Preis Auserhaus"
    ws["E1"].value = "MWSt"
    ws["F1"].value = "MWSt Auserhaus"

    tmp_artikel_Sparten = []
    color = 0x77bb22

    # gehe über alle Artikel in der plist
    for i in range(len(artikel_Data)):
        (id_entry, artikel_entry, preis_entry, preisah_entry, mwst_entry, mwstah_entry, sparte, sparte_id, new_artikle) = artikel_Data[i]

        if sparte not in tmp_artikel_Sparten:
                tmp_artikel_Sparten.insert(sparte_id, sparte)
                color = color + 12255
                

        # wählt die Zelle aus wo der jeweilige eintrag rein soll
        id = ws["A" + str(i + 2)]
        artikel = ws["B" + str(i + 2)]
        preis = ws["C" + str(i + 2)]
        preisah = ws["D" + str(i + 2)]
        mwst = ws["E" + str(i + 2)]
        mwstah = ws["F" + str(i + 2)]

        # Werte aus der .plist in die jeweiligen Zellen kopiert
        # gegfalls muss das komma mit einem punkt ausgetauscht werden!
        id.value = id_entry
        artikel.value = artikel_entry
        preis.value = float(preis_entry.replace(",", "."))
        preisah.value = float(preisah_entry.replace(",", "."))
        mwst_in_percent = float(mwst_entry.replace(",", ".")) / 100
        mwst.value = mwst_in_percent
        mwstah_in_percent = float(mwstah_entry.replace(",", ".")) / 100
        mwstah.value = mwstah_in_percent

        # formatierung der % felder und setzt die IDs auf Kursiv
        id.font = Font(italic=True)
        id.fill = PatternFill(start_color=format(color, "x"), end_color=format(color, "x"), fill_type='solid')
        mwst.style = "Percent"
        mwstah.style = "Percent"

        # Passt die Breite der Reihen an für better looks :)
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max(
                        (dims.get(cell.column_letter, 0), len(str(cell.value)))
                    )
        for col, value in dims.items():
            ws.column_dimensions[col].width = value

    # die Exccel Datei auf Disk speichern
    try:
        wb.save("exoprtArtikel.xlsx")
    except:
        messagebox.showerror(message="Fehler beim erstellen der Excel Datei")

    messagebox.showinfo(message="Excel Datei ersetllt.", title="Information")
    
def preis_to_preis_ah():
    if not artikel_Data:
        return
    
    for i in range(len(artikel_Data)):
        (id, name, preis, preisah, mwst, mwstah, sparte, sparte_id, new_item) = artikel_Data[i]

        artikel_Data[i] = (id, name, preis, preis, mwst, mwstah, sparte, sparte_id, new_item)


    poulate_ui()

def mwst_to_mwst_ah():
    if not artikel_Data:
        return
    
    for i in range(len(artikel_Data)):
        (id, name, preis, preisah, mwst, mwstah, sparte, sparte_id, new_item) = artikel_Data[i]

        artikel_Data[i] = (id, name, preis, preisah, mwst, mwst, sparte, sparte_id, new_item)
    
    poulate_ui()

def update_all_mwst():
    if not artikel_Data:
        return
    
    new_mwst = simpledialog.askfloat("MWSt", "Für alle MWSt eingeben:", initialvalue=19.00)
    
    new_mwst = str(new_mwst).replace(".", ",")

    
    for i in range(len(artikel_Data)):
        (id, name, preis, preisah, mwst, mwstah, sparte, sparte_id, new_item) = artikel_Data[i]

        artikel_Data[i] = (id, name, preis, preisah, new_mwst, mwstah, sparte, sparte_id, new_item)
    
    poulate_ui()

import_button = ttk.Button(root, text="Import Artikel", command=select_File)
import_button.pack()
save_button = ttk.Button(root, text="Speichern", command=save_Data)
save_button.pack()
export_button = ttk.Button(root, text="Export Artikel", command=save_File)
export_button.pack()
excel_button = ttk.Button(root, text="Export Excel", command=save_to_excel)
excel_button.pack()
new_artikle = ttk.Button(root, text="Neuen Artikel", command=create_new_artikle_widget)
new_artikle.pack()
preisispreisah = ttk.Button(root, text="Preis = Preis AH", command=preis_to_preis_ah)
preisispreisah.pack()
mwstismwstah = ttk.Button(root, text="MWSt = MWSt AH", command=mwst_to_mwst_ah)
mwstismwstah.pack()
mwstismwstah = ttk.Button(root, text="Alle MWSt = ", command=update_all_mwst)
mwstismwstah.pack()

messagebox.showinfo(message="Dieses Programm wird so bereitgestellt, wie es ist. \nIch übernehme keine Verantwortung für Probleme, die auftreten können! \nBitte überprüfe das Endresultat vor dem Einspielen in die Kasse.", title="Information!")

poulate_info()

root.mainloop()